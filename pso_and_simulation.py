import os
import sys
import random
import logging
import warnings
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
from sklearn.model_selection import KFold
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

current_dir = os.path.dirname(os.path.abspath(__file__))
mem_mem_dir = os.path.join(current_dir, 'MTC_MEM')
sys.path.append(mem_mem_dir)

from model import get_model
from mem_simulator import Simulation
from read import ReadData
from adaptive_pso import  StepwisePSO
from result_utils import create_results_directory, load_config, over_sample_high_eta, update_membrane_json

warnings.filterwarnings("ignore")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("app.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)


def append_df_to_excel_direct(filepath, df_new, sheet_name='Sheet1'):
    """追加或创建Excel表，并写入DataFrame"""
    if not os.path.isfile(filepath):
        try:
            df_new.to_excel(filepath, index=False, sheet_name=sheet_name)
        except Exception as e:
            logging.error(f"Failed to create Excel file: {e}")
        return
    try:
        workbook = load_workbook(filepath)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in dataframe_to_rows(df_new, index=False, header=False):
                sheet.append(row)
        else:
            sheet = workbook.create_sheet(sheet_name)
            for r in dataframe_to_rows(df_new, index=False, header=True):
                sheet.append(r)
        workbook.save(filepath)
    except Exception as e:
        logging.error(f"Failed to append data to Excel: {e}")


def simulate_once(params):
    """Execute a single simulation and return results as a DataFrame."""
    try:
        json_path = "./MTC_MEM/in_mem.json"
        update_membrane_json(params, json_path)
        in_data = ReadData(kn_model='BU')
        in_data.feed_para['condition']['T'][0] = params.get('T0', 0.0)
        in_data.feed_para['condition']['P'][0] = round(params.get('P0', 0.0), 2)


        sim = Simulation(
            in_data.reactor_data.iloc[0],
            in_data.chem,
            in_data.feed_data.iloc[0],
            in_data.mem_data.iloc[0],
            eos=1,
            drop=1,
          #  dist_opt_config=dist_opt_config
        )
        return sim.sim(save_profile=0, loop='indirect', rtol=0.001, r_target=None)

    except Exception as e:
        logging.error(f"Simulation failed: {e}")
        return None


def objective_function_1(params, model, scaler_X, scaler_Y, config):
    """Convert PSO parameters -> Model prediction -> Return negative objective value."""
    var_keys = list(config['variable_parameters'].keys())
    param_dict = dict(zip(var_keys, params))
    for k, v in param_dict.items():
        try:
            param_dict[k] = float(v)
        except ValueError:
            param_dict[k] = 0.0
    input_cols = config.get('input_columns', [])
    if not input_cols:
        return np.inf
    try:
        feature_vector = [param_dict[col] for col in input_cols]
        X_scaled = scaler_X.transform([feature_vector])
        y_scaled = model.predict(X_scaled)
        y_pred = scaler_Y.inverse_transform(y_scaled.reshape(-1, 1)).ravel()[0]
    except Exception:
        return np.inf
    if y_pred is None:
        return np.inf
    target_eta = config.get('target_eta_threshold', 0.9)
    return ((target_eta - y_pred) ** 2 * 10) if (y_pred < target_eta) else -y_pred


def iterative_main():
    """Main process: Read config -> Train model -> PSO -> Simulate -> Iterative update."""
    config = load_config()
    seed = config.get('seed', 42)
    np.random.seed(seed)
    random.seed(seed)

    fixed_params = config.get('fixed_parameters', {})
    var_params = config.get('variable_parameters', {})
    input_cols = [col.lower() for col in config.get('input_columns', [])]
    target_cols = [col.lower() for col in config.get('output_columns', [])]
    run_folder = create_results_directory()


    DATA_PATH = config.get('data_path', './data/raw_data.xlsx')
    opt_cfg = config.get('optimization', {})
    qmh1_option = opt_cfg.get('qmh1_option', 1)
    th1 = opt_cfg.get('th1', 0.5)
    try:
        df = pd.read_excel(DATA_PATH, engine='openpyxl')
        df.columns = df.columns.str.strip().str.lower()
        df['qmh1'] = df['qmh1'].astype(int)
        df_filtered = df[(df['qmh1'] == qmh1_option) & (df['eta_mod_heat'] >= th1)]
        df_filtered = df_filtered.dropna(subset=input_cols + target_cols)
    except Exception as e:
        logging.error(f"Data loading error: {e}")
        return

    scaler_X = StandardScaler()
    scaler_Y = StandardScaler()


    def train_model(df_data, model_name='mlp', input_cols=None, target_cols=None):
        """
        完整的模型训练流程

        Args:
            df_data: 输入数据框
            model_name: 模型名称，默认为'mlp'
            input_cols: 输入特征列名列表
            target_cols: 目标变量列名列表

        Returns:
            tuple: (训练好的模型, 性能指标字典)
        """
        if df_data.empty:
            logging.warning("Empty input dataframe")
            return None, None


        df_data = over_sample_high_eta(df_data, eta_threshold=0.89, multiple=5)

        X = df_data[input_cols].values
        Y_unscaled = df_data[target_cols].values.ravel()

        if len(X) == 0:
            logging.warning("No features extracted")
            return None, None


        X_scaled = scaler_X.fit_transform(X)
        Y_scaled = scaler_Y.fit_transform(Y_unscaled.reshape(-1, 1)).ravel()


        sample_weight_full = np.ones(X.shape[0], dtype=np.float32)
        sample_weight_full[Y_unscaled > 0.8] = 5.0 

        if model_name == 'xgboost_optuna':

            final_model = get_model(
                model_name,
                random_state=42,
                X_train=X_scaled,
                y_train=Y_scaled,
                sample_weight=sample_weight_full
            )
            X_train, X_val, y_train, y_val, sw_train, sw_val = train_test_split(
                X_scaled, Y_scaled, sample_weight_full,
                test_size=0.2,
                random_state=42
            )

            try:
                final_model.set_params(early_stopping_rounds=50)  # 通过 set_params 设置早停
                final_model.fit(
                    X_train,
                    y_train,
                    sample_weight=sw_train,
                    eval_set=[(X_val, y_val)],
                    eval_metric='rmse',
                    verbose=False
                )

                # 获取最佳迭代次数
                best_iteration = final_model.best_iteration

                # 用全量数据重新训练，使用找到的最佳迭代次数
                final_model.set_params(early_stopping_rounds=None)  # 移除早停
                final_model.set_params(n_estimators=best_iteration)  # 使用找到的最佳迭代次数

                final_model.fit(
                    X_scaled,
                    Y_scaled,
                    sample_weight=sample_weight_full,
                    verbose=False
                )
            except TypeError:
                logging.warning("Model doesn't support sample weights, training without weights")
                final_model.fit(X_scaled, Y_scaled)

        else:

            base_model = get_model(model_name)

            kf = KFold(n_splits=5, shuffle=True, random_state=42)
            cv_scores = []

            for train_idx, val_idx in kf.split(X_scaled):
                X_train, X_val = X_scaled[train_idx], X_scaled[val_idx]
                y_train, y_val = Y_scaled[train_idx], Y_scaled[val_idx]
                sw_train = sample_weight_full[train_idx]

                try:

                    base_model.fit(X_train, y_train, sample_weight=sw_train)
                except TypeError:

                    base_model.fit(X_train, y_train)
                except Exception as e:
                    logging.error(f"Error in CV fit: {e}")
                    return None, None

                val_pred = base_model.predict(X_val)
                cv_scores.append(mean_squared_error(y_val, val_pred, squared=False))

            logging.info(f"Cross-validation RMSE: {np.mean(cv_scores):.4f} ± {np.std(cv_scores):.4f}")

            final_model = get_model(model_name)
            try:
                final_model.fit(X_scaled, Y_scaled, sample_weight=sample_weight_full)
            except TypeError:
                final_model.fit(X_scaled, Y_scaled)

        pred = final_model.predict(X_scaled)

        metrics = {
            'r2': r2_score(Y_scaled, pred),
            'rmse': np.sqrt(mean_squared_error(Y_scaled, pred)),
            'mae': mean_absolute_error(Y_scaled, pred)
        }

        final_model.scaler_X = scaler_X
        final_model.scaler_Y = scaler_Y

        logging.info(f"Training metrics: {metrics}")
        return final_model, metrics

    model_name = config.get('model', {}).get('model_name', 'mlp')
    model, init_perf = train_model(df_filtered, model_name, input_cols, target_cols)
    if model is None:
        logging.error("Initial model training failed.")
        return

    bounds_lower = np.array([p['lower'] for p in var_params.values()])
    bounds_upper = np.array([p['upper'] for p in var_params.values()])
    n_iterations = opt_cfg.get('n_iterations', 100)
    os.makedirs(run_folder, exist_ok=True)

    new_rows_path = os.path.join(run_folder, 'new_rows.xlsx')

    records = [{'Iteration': 0, **init_perf}]
    pso_cfg = config.get('pso_parameters', {})

    for it in range(n_iterations):
        logging.info(f"=== Iteration {it + 1}/{n_iterations} ===")

        pso = StepwisePSO(
            model=model,
            scaler_X=model.scaler_X,
            scaler_Y=model.scaler_Y,
            input_cols=input_cols,
            bounds_lower=bounds_lower,
            bounds_upper=bounds_upper,
            num_particles=int(pso_cfg.get('num_particles', 30)),
            max_iter=int(pso_cfg.get('max_iter', 100)),
            early_stopping_rounds=pso_cfg.get('early_stopping_rounds', 10),
            early_stopping_tolerance=pso_cfg.get('early_stopping_tolerance', 1e-4),
            w=0.7,
            c1=1.5, 
            c2=1.5 
        )

        base_run = os.path.join(run_folder, f"iteration_{it + 1}")
        os.makedirs(base_run, exist_ok=True)

        num_runs = pso_cfg.get('num_runs', 5)
        all_results = []

        for run_idx in range(num_runs):
            sub_run = os.path.join(base_run, f"run_{run_idx + 1}")
            os.makedirs(sub_run, exist_ok=True)

            try:

                res = pso.two_stage_optimize(
                    stage1_dims=[4, 5],
                    stage1_fixed={ 
                        0: 1e-6, 
                        1: 100.0, 
                        2: 0.2,
                        3: 0.0 
                    },
                    stage2_dims=[0, 1, 2], 
                    stage2_fixed={ 
                        3: 0.0,
                        4: None, 
                        5: None  
                    },
                    stage1_iter=int(pso_cfg.get('max_iter', 100) // 2), 
                    stage2_iter=int(pso_cfg.get('max_iter', 100) // 2), 
                    folder_stage1=os.path.join(sub_run, "stage1_TP"),
                    folder_stage2=os.path.join(sub_run, "stage2_3params"),
                    local_search_maxiter=50
                )

                logging.info(f"Run {run_idx + 1} finished, best_eta: {res['final_best_eta']:.4f}")
                all_results.append(res)
            except Exception as e:
                logging.error(f"PSO run {run_idx + 1} failed: {e}")


        best_idx = np.argmin([res['stage2_result']['best_cost'] for res in all_results])
        best_res = all_results[best_idx]
        best_params = best_res['final_best_params']
        best_eta = best_res['final_best_eta']
        logging.info(f"Iteration {it + 1}: Best solution: {best_params}, best_eta: {best_eta:.4f}")


        p_dict = dict(zip(var_params.keys(), best_params))
        p_dict.update(fixed_params)
        for k,v in p_dict.items():
            try:
                p_dict[k] = float(v)
            except:
                p_dict[k] = 0.0

        res = simulate_once(
            p_dict
        )
        if not (isinstance(res, pd.DataFrame) and not res.empty):
            continue

        final_data = res.iloc[-1].to_dict()

        new_row = {
            'per_h2o1': p_dict['per_H2O1'],
            'per_s1': p_dict['per_S1'],
            'fp1': p_dict['Fp1'],
            'fp2': p_dict['Fp2'],
            't0': p_dict['T0'],
            'p0': p_dict['P0'],
            'work': final_data.get('work', np.nan),
            'q_per': final_data.get('q_per', np.nan),
            'q_exergy': final_data.get('q_exergy', np.nan),
            'eta_mod_heat': final_data.get('eta_mod_heat', np.nan),
            'y_ch3oh_c': final_data.get('y_CH3OH_C', np.nan),
            'p_conv': final_data.get('p_conv', np.nan),
            'pso_eta_pred': best_eta
        }
        nr_df = pd.DataFrame([new_row]).astype(float, errors='ignore')
        append_df_to_excel_direct(new_rows_path, nr_df)

        df_filtered = pd.concat([df_filtered, nr_df], ignore_index=True)
        model, cur_perf = train_model(df_filtered, model_name, input_cols, target_cols)
        if model is None:
            break
        records.append({'Iteration': it+1, **cur_perf})


        if it > 0:
            impr = records[-1]['r2'] - records[-2]['r2']
            if impr < 0.001:
                pso.num_particles = min(pso.num_particles + 50, 1000)
                pso.max_iter += 50


    pf_df = pd.DataFrame(records)
    pf_path = os.path.join(run_folder, 'model_performance.xlsx')
    pf_df.to_excel(pf_path, index=False)

    plt.figure(figsize=(10, 6))
    plt.subplot(3,1,1)
    plt.plot(pf_df['Iteration'], pf_df['r2'], 'o-')
    plt.axhline(y=0.9, color='r', linestyle='--')
    plt.title('Performance Over Iterations')
    plt.ylabel('R²'); plt.grid(True)

    plt.subplot(3,1,2)
    plt.plot(pf_df['Iteration'], pf_df['rmse'], 'o-g')
    plt.ylabel('RMSE'); plt.grid(True)

    plt.subplot(3,1,3)
    plt.plot(pf_df['Iteration'], pf_df['mae'], 'o-m')
    plt.xlabel('Iteration')
    plt.ylabel('MAE'); plt.grid(True)

    plt.tight_layout()
    plt.savefig(os.path.join(run_folder, 'model_performance.png'))
    plt.close()


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    iterative_main()
